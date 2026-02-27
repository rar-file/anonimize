"""Excel file format handler.

This module provides support for reading and writing Excel files
(.xlsx, .xls) with streaming capabilities.
"""

from pathlib import Path
from typing import Any, BinaryIO, Dict, Iterator, List, Optional, Union
import logging

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    Workbook = None
    load_workbook = None
    dataframe_to_rows = None
    XLTable = None
    TableStyleInfo = None

from anonimize.formats.base import (
    BaseFormatHandler,
    FormatConfig,
    FileStats,
    StreamingWriter,
    register_handler,
)

logger = logging.getLogger(__name__)


class ExcelStreamingWriter(StreamingWriter):
    """Streaming writer for Excel files.
    
    Uses openpyxl to write batches incrementally.
    Note: Excel files must be written in memory and saved at the end
    due to the nature of the format.
    """
    
    def __init__(
        self,
        destination: Union[str, Path, BinaryIO],
        config: FormatConfig,
        schema: Optional[Dict[str, str]] = None,
        sheet_name: str = "Sheet1",
    ):
        """Initialize the Excel streaming writer."""
        super().__init__(destination, config, schema)
        
        self.sheet_name = sheet_name
        self._workbook = Workbook()
        self._worksheet = self._workbook.active
        self._worksheet.title = sheet_name
        self._headers_written = False
        self._headers = []
    
    def write_batch(self, batch: List[Dict[str, Any]]) -> int:
        """Write a batch of rows to the Excel file."""
        self._check_closed()
        
        if not batch:
            return 0
        
        # Write headers on first batch
        if not self._headers_written:
            self._headers = list(batch[0].keys())
            self._worksheet.append(self._headers)
            self._headers_written = True
        
        # Write data rows
        for row in batch:
            row_values = [row.get(col) for col in self._headers]
            self._worksheet.append(row_values)
        
        self._rows_written += len(batch)
        
        return len(batch)
    
    def close(self) -> None:
        """Close the writer and save the Excel file."""
        if self._closed:
            return
        
        # Auto-adjust column widths
        if self._headers_written:
            for column in self._worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                self._worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        if isinstance(self.destination, (str, Path)):
            self._workbook.save(self.destination)
        else:
            # Assume file-like object
            self._workbook.save(self.destination)
        
        self._workbook.close()
        self._closed = True
        
        logger.debug(f"Excel writer closed. Rows written: {self._rows_written}")


class ExcelHandler(BaseFormatHandler):
    """Handler for Excel file format (.xlsx, .xls).
    
    This handler provides support for reading and writing Excel files
    with the following features:
    - Multiple sheet support
    - Column type inference
    - Table formatting
    - Auto column width
    - Streaming writes
    
    Example:
        >>> from anonimize.formats.excel import ExcelHandler
        >>> handler = ExcelHandler()
        >>> # Read entire file
        >>> data = handler.read("data.xlsx")
        >>> # Read specific sheet
        >>> data = handler.read("data.xlsx", sheet_name="Users")
        >>> # Write data
        >>> handler.write("output.xlsx", data, sheet_name="Users")
        >>> # Write multiple sheets
        >>> handler.write_multi(
        ...     "output.xlsx",
        ...     {"Users": users_data, "Orders": orders_data}
        ... )
    """
    
    def __init__(self, config: Optional[FormatConfig] = None):
        """Initialize the Excel handler.
        
        Args:
            config: Format configuration.
        
        Raises:
            ImportError: If required dependencies are not installed.
        """
        if not PANDAS_AVAILABLE:
            raise ImportError(
                "pandas is required for Excel support. "
                "Install it with: pip install pandas"
            )
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install it with: pip install openpyxl"
            )
        
        super().__init__(config)
    
    @property
    def supported_extensions(self) -> List[str]:
        """Return supported file extensions."""
        return [".xlsx", ".xls"]
    
    def read(
        self,
        source: Union[str, Path, BinaryIO],
        columns: Optional[List[str]] = None,
        **kwargs
    ) -> List[Dict[str, Any]]:
        """Read data from an Excel file.
        
        Args:
            source: File path or file-like object.
            columns: Optional list of columns to read.
            **kwargs: Additional options:
                - sheet_name: Sheet to read (default: 0/first sheet)
                - header: Row to use as header (default: 0)
                - skiprows: Number of rows to skip
                - nrows: Maximum number of rows to read
        
        Returns:
            List of row dictionaries.
        """
        sheet_name = kwargs.get("sheet_name", 0)
        header = kwargs.get("header", 0)
        skiprows = kwargs.get("skiprows", None)
        nrows = kwargs.get("nrows", None)
        
        df = pd.read_excel(
            source,
            sheet_name=sheet_name,
            header=header,
            skiprows=skiprows,
            nrows=nrows,
            usecols=columns,
            engine="openpyxl",
        )
        
        # Convert to list of dicts
        result = df.replace({pd.NA: None, float('nan'): None}).to_dict("records")
        
        self._stats.rows_read += len(result)
        self._stats.columns = list(df.columns)
        
        return result
    
    def read_streaming(
        self,
        source: Union[str, Path, BinaryIO],
        columns: Optional[List[str]] = None,
        batch_size: Optional[int] = None,
        **kwargs
    ) -> Iterator[List[Dict[str, Any]]]:
        """Read Excel file in batches (streaming).
        
        Note: Excel files don't support true streaming, so this reads
        in chunks using pandas.
        
        Args:
            source: File path or file-like object.
            columns: Optional list of columns to read.
            batch_size: Number of rows per batch.
            **kwargs: Additional options (same as read).
        
        Yields:
            Batches of row dictionaries.
        """
        batch_size = batch_size or self.config.batch_size
        
        # Read entire file (Excel doesn't support chunking natively)
        df = pd.read_excel(
            source,
            sheet_name=kwargs.get("sheet_name", 0),
            header=kwargs.get("header", 0),
            usecols=columns,
            engine="openpyxl",
        )
        
        self._stats.columns = list(df.columns)
        
        # Yield in batches
        total_rows = len(df)
        for start in range(0, total_rows, batch_size):
            end = min(start + batch_size, total_rows)
            batch_df = df.iloc[start:end]
            batch = batch_df.replace({pd.NA: None, float('nan'): None}).to_dict("records")
            self._stats.rows_read += len(batch)
            yield batch
    
    def write(
        self,
        destination: Union[str, Path, BinaryIO],
        data: List[Dict[str, Any]],
        schema: Optional[Dict[str, str]] = None,
        **kwargs
    ) -> FileStats:
        """Write data to an Excel file.
        
        Args:
            destination: File path or file-like object.
            data: List of row dictionaries.
            schema: Optional schema (used for column ordering).
            **kwargs: Additional options:
                - sheet_name: Sheet name (default: "Sheet1")
                - index: Include index column (default: False)
                - as_table: Format as Excel table (default: True)
                - table_style: Table style name
        
        Returns:
            File statistics.
        """
        if not data:
            # Create empty workbook
            wb = Workbook()
            wb.save(destination)
            return FileStats(rows_written=0)
        
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        index = kwargs.get("index", False)
        as_table = kwargs.get("as_table", True)
        table_style = kwargs.get("table_style", "TableStyleMedium9")
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Reorder columns if schema provided
        if schema:
            ordered_cols = [col for col in schema.keys() if col in df.columns]
            remaining_cols = [col for col in df.columns if col not in schema]
            df = df[ordered_cols + remaining_cols]
        
        # Write using openpyxl for more control
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        headers = list(df.columns)
        ws.append(headers)
        
        # Write data
        for row in df.itertuples(index=False):
            # Replace NaN with None
            row_values = [None if pd.isna(val) else val for val in row]
            ws.append(row_values)
        
        # Format as table if requested
        if as_table and len(data) > 0:
            table_ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(data) + 1}"
            tab = XLTable(displayName=sheet_name.replace(" ", "_"), ref=table_ref)
            
            style = TableStyleInfo(
                name=table_style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(destination)
        wb.close()
        
        self._stats.rows_written += len(data)
        self._stats.columns = headers
        
        return self._stats
    
    def write_streaming(
        self,
        destination: Union[str, Path, BinaryIO],
        schema: Optional[Dict[str, str]] = None,
        **kwargs
    ) -> ExcelStreamingWriter:
        """Create a streaming writer for Excel files.
        
        Args:
            destination: File path or file-like object.
            schema: Optional schema dictionary.
            **kwargs: Additional options:
                - sheet_name: Sheet name (default: "Sheet1")
        
        Returns:
            ExcelStreamingWriter instance.
        """
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        return ExcelStreamingWriter(destination, self.config, schema, sheet_name)
    
    def write_multi(
        self,
        destination: Union[str, Path, BinaryIO],
        data_dict: Dict[str, List[Dict[str, Any]]],
        **kwargs
    ) -> FileStats:
        """Write multiple sheets to an Excel file.
        
        Args:
            destination: File path or file-like object.
            data_dict: Dictionary mapping sheet names to data lists.
            **kwargs: Additional options passed to write.
        
        Returns:
            File statistics.
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        total_rows = 0
        all_columns = set()
        
        for sheet_name, data in data_dict.items():
            ws = wb.create_sheet(title=sheet_name)
            
            if not data:
                continue
            
            df = pd.DataFrame(data)
            headers = list(df.columns)
            all_columns.update(headers)
            
            # Write headers
            ws.append(headers)
            
            # Write data
            for row in df.itertuples(index=False):
                row_values = [None if pd.isna(val) else val for val in row]
                ws.append(row_values)
            
            total_rows += len(data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(destination)
        wb.close()
        
        self._stats.rows_written += total_rows
        self._stats.columns = list(all_columns)
        
        return self._stats
    
    def get_schema(self, source: Union[str, Path, BinaryIO]) -> Dict[str, str]:
        """Get the schema of an Excel file.
        
        Args:
            source: File path or file-like object.
        
        Returns:
            Dictionary mapping column names to pandas types.
        """
        df = pd.read_excel(
            source,
            sheet_name=0,
            header=0,
            nrows=1,
            engine="openpyxl",
        )
        
        return {col: str(dtype) for col, dtype in df.dtypes.items()}
    
    def get_sheet_names(self, source: Union[str, Path, BinaryIO]) -> List[str]:
        """Get list of sheet names from an Excel file.
        
        Args:
            source: File path or file-like object.
        
        Returns:
            List of sheet names.
        """
        wb = load_workbook(source, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names


# Register handler
if PANDAS_AVAILABLE and OPENPYXL_AVAILABLE:
    register_handler(ExcelHandler())
